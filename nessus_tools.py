import os
import glob
import pandas as pd
import xml.etree.ElementTree as ET
import zipfile
from pathlib import Path
from typing import List
import tkinter as tk
from tkinter import filedialog
import re
import shutil
import datetime
import time
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


class NessusParser:
    '''
    Parses a Nessus XML file and extracts relevant information.
    This class is designed to handle Nessus XML files, extracting findings
    and exporting them to an Excel file.
    '''
    def __init__(self, filepath):
        self.filepath = filepath
        self.tree = None
        self.root = None
        self._load()

    def _load(self):
        '''
        Loads the XML file and parses it into an ElementTree object.
        '''
        try:
            self.tree = ET.parse(self.filepath)
            self.root = self.tree.getroot()
            Logger.log(f"✅  Succesfully loaded XML file: {self.filepath}")
        except Exception as e:
            Logger.log(f"❌ Failed to load XML file {self.filepath}: {e}")
            raise RuntimeError(f"Failed to load XML file {self.filepath}: {e}")

    @staticmethod
    def parse_descript_block(text: str) -> dict:
        '''
        Parse a Nessus description block into STIG, FINDING, Actual Value, Pasteable, and Short Desc columns.
        '''
        result = {
            "STIG": "",
            "FINDING": "",
            "Actual Value": "",
            "Pasteable": "",
            "Short Desc": "",
            "Solution": ""
        }

        if not text:
            Logger.log("⚠️ Empty description block encountered")
            return result
        
        # Normalize line breaks
        lines = text.splitlines()
        clean_text = "\n".join([l.strip() for l in lines if l.strip()])
        Logger.log(f"🔍 Parsing description block (first 80 chars): {clean_text[:80]}...")

        # --- Extract STIG UD + Finding ---
        # Example: "CASA-FW-000200 - The Cisco ASA must be configured ... : [FAILED]"
        match = re.search(
            r'^"([A-Z0-9\-]+)\s*-\s*(.+?)"\s*: \[FAILED\]',
            clean_text, 
            re.DOTALL
            )
        
        if match:
            stig, finding       = match.groups()
            result["STIG"]      = stig.strip()
            result["FINDING"]   = finding.strip()
            result["Pasteable"] = f"{stig} - {finding}"
            Logger.log(f"✅ Regex matched STIG={stig}, FINDING starts with={finding[:40]}...")
        else:
            Logger.log("❌ Regex failed for description block")
            Logger.log(f"Full description block:\n{clean_text}")

        # --- Extract Actual Value ---
        m_actual = re.search(r"Actual Value:\s*(.+)", clean_text, re.IGNORECASE | re.DOTALL)
        if m_actual:
            # Grab until next section or end
            actual_value = m_actual.group(1).strip()
            result["Actual Value"] = actual_value

        # --- Extract Short Desc ---
        # Everything between the finding line and "Solution:" is the short description
        short_match = re.search(r"\[FAILED\](.*)Solution:", clean_text, re.DOTALL | re.IGNORECASE)
        if short_match:
            result["Short Desc"] = short_match.group(1).strip()

        # --- Extract Solution ---
        # Everything between "Solution:" and the next section is the solution
        solution_match = re.search(r"Solution:\s*(.*)\n", clean_text, re.DOTALL | re.IGNORECASE)
        if solution_match:
            result["Solution"] = solution_match.group(1).strip()

        return result

    def get_cat_findings(self, cat_lvls=("II",)):
        '''
        Extract findings that match any of the requested CAT levels,
        and include the host associated with the finding.
        Args:
            cat_lvls (tuple|list): e.g., ("II",) or ("I", "II", "III")   
        '''
        file_start_time = time.time() # ⏱ start timer for per-file processing time

        print(f"📂 Parsing Nessus file: {os.path.splitext(os.path.basename(self.filepath))[0]}")
        print(f"⚙️  Extracting findings for CAT levels: {cat_lvls}")
        Logger.log(f"⚙️ Extracting CAT findings for levels: {cat_lvls} from 📂 {os.path.splitext(os.path.basename(self.filepath))[0]}")

        findings = []
        cat_lvls = [f"CAT|{lvl.upper()}" for lvl in cat_lvls]

        for report in self.root.findall(".//Report"):
            for host in report.findall(".//ReportHost"):
                hostname = host.get("name", "UNKNOWN")
                Logger.log(f"📡 Processing host: {hostname}")

                for item in host.findall(".//ReportItem"):
                    Logger.log(f"📑 Processing ReportItem...")
                    plugin_id = item.get("pluginID")
                    severity = item.get("severity")
                    plugin_name = item.get("pluginName")
                    description = item.findtext("description", default="").strip()

                    # ----- Compliance findings ---
                    compliance_ref = (item.findtext("{*}compliance-reference") or "").strip()
                    compliance_result = (item.findtext("{*}compliance-result") or "").strip()

                    # Only process FAILED compliance results
                    if compliance_result != "FAILED":
                        #Logger.log(f"Compliance check {compliance_result} - going to next item.") #NOT NEEDED
                        continue #skips to next item

                    # Extract actual CAT level from compliance ref
                    actual_cat = ""
                    for part in compliance_ref.split(","):
                        kv = [p.strip() for p in part.split("|", 1)]
                        if len(kv) == 2 and kv[0].upper() == "CAT":
                            actual_cat = kv[1].upper()
                            break

                    # Skip if it's not one of the requested CAT levels
                    if f"CAT|{actual_cat}" not in cat_lvls:
                        Logger.log(f" Skipped finding (CAT {actual_cat} not in requested levels) Host={hostname}, Plugin={plugin_id}")
                        continue #skips to next item

                    # Check if any requested CAT level is present in the compliance result
                    matched = False
                    for cat_lvl in cat_lvls:
                        if cat_lvl in compliance_ref.upper():
                            # ✅ Only parse description block when we know it’s relevant
                            parsed = NessusParser.parse_descript_block(description)

                            matched = True
                            findings.append({
                                "Hostname": hostname,
                                "Plugin ID": plugin_id,
                                "CAT": actual_cat,
                                "Severity": severity,
                                "Result": compliance_result,
                                "STIG": parsed["STIG"],
                                "FINDING": parsed["FINDING"],
                                "Actual Value": parsed["Actual Value"],
                                "Short Desc": parsed["Short Desc"],
                                "Plugin Name": plugin_name,
                                #"Description": description.strip(),
                                "Pasteable": parsed["Pasteable"],
                                "Compliance Reference": compliance_ref.strip(),
                            })
                            Logger.log(f"✅ Added finding from {hostname} (Plugin {plugin_id}, {actual_cat})")
                            break

                        if not matched:
                            Logger.log(f"ℹ️ Skipped finding (no CAT match) Host={hostname}, Plugin={plugin_id}")

      
        file_end_time = time.time() # ⏱ end timer
        elapsed = file_end_time - file_start_time
        minutes, seconds = divmod(int(elapsed), 60)

        if findings:
            Logger.log(f"📊 Total findings extracted: {len(findings)} in {minutes}:{seconds}")
            return pd.DataFrame(findings), (minutes, seconds)
        else:
            df_empty = pd.DataFrame(columns=[
                "Hostname",
                "Plugin ID",
                "CAT",
                "Severity",
                "Result",
                "STIG",
                "FINDING",
                "Actual Value",
                "Short Desc",
                "Plugin Name",
                "Description",
                "Pasteable",
                "Cross References",
                "Compliance Reference",
                "Compliance Result"
            ])
            return df_empty, (minutes, seconds)

    
class NessusToExcelExporter:
    '''Exports findings to an Excel file.'''
    def __init__(self, input_folder, output_file, cat_lvls=("II",)):
        self.input_folder = input_folder
        self.output_file = output_file
        self.cat_lvls = cat_lvls
        self.files = glob.glob(os.path.join(input_folder, '*.nessus'))

    def run(self):
        '''Process all files and export CAT:II findings to Excel.'''
        if not self.files:
            print(f"No Nessus files found in {self.input_folder}")
            Logger.log(f"No Nessus files found in {self.input_folder}")
            return
        
        with pd.ExcelWriter(self.output_file, engine='openpyxl') as writer:
            any_written = False
            print(f"⚙️  Processing Nessus files in {self.input_folder}...")
            Logger.log(f"⚙️  Processing Nessus files in {self.input_folder}...")

            for filepath in self.files:
                try:
                    parser = NessusParser(filepath)
                    df, (mins, secs) = parser.get_cat_findings(cat_lvls=self.cat_lvls)
                    fn = os.path.splitext(os.path.basename(filepath))[0]
                    
                    if not df.empty:
                        # Add a blank "Comments" column
                        df["Comments"] = ""
                        # Reorder to make sure Comments is at the end
                        df = df[[*df.columns.drop("Comments"), "Comments"]]

                        # Use filename as sheet name, limited to 31 characters
                        sheet_name = os.path.splitext(os.path.basename(filepath))[0][:31]
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                        any_written = True
                        print(f"✅ Processed {sheet_name}, found {len(df)} findings (CAT {','.join(self.cat_lvls)}) in {mins}:{secs}.")
                        Logger.log(f"✅ Processed {filepath}, found {len(df)} findings (CAT {','.join(self.cat_lvls)}) in {mins}:{secs}.")
                    else:
                        print(f"ℹ️  No CAT {','.join(self.cat_lvls)} findings in {fn}")
                        Logger.log(f"ℹ️  No CAT {','.join(self.cat_lvls)} findings in {filepath}")
                except Exception as e:
                    print(f"❌ Error processing {filepath}: {e}")
                    Logger.log(f"❌ Error processing {filepath}: {e}")
                    continue
            if not any_written:
                placeholder_df = pd.DataFrame(
                    {
                        "Message": [
                            f"No CAT {','.join(self.cat_lvls)} findings found in any file."
                        ]
                    }
                )
                placeholder_df.to_excel(writer, sheet_name="No_Findings", index=False)
                print("⚠️  No findings found in any file. Wrote placeholder sheet")
                Logger.log("⚠️  No findings found in any file. Wrote placeholder sheet")

        # 🔹 Post-process with openpyxl for Wrap Text
        wb = load_workbook(self.output_file)
        for ws in wb.worksheets:
            # Freeze top row
            ws.freeze_panes = "A2"

            # Format head row: bold + centered
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Auto column width + wrap text
            for col_idx, col_cells in enumerate(ws.columns, 1):
                max_length = 0
                col_letter = get_column_letter(col_idx)
                for cell in col_cells:
                    cell.alignment = Alignment(wrap_text=True)
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min((max_length + 2), 60) # cap width so it's not crazy wide
                ws.column_dimensions[col_letter].width = adjusted_width
        wb.save(self.output_file)


        print(f"\n🎉 Finished! Findings (CAT {','.join(self.cat_lvls)}) saved in {self.output_file}")
        Logger.log(f"🎉 Finished! Findings (CAT {','.join(self.cat_lvls)}) saved in {self.output_file}")

class NessusExtractor:
    '''
    - Extracts .nessus files from Tenable Security Center ZIP scan downloads
    into a single folder.
    - Renames them based on the ZIP filename (removing 'PAAN_' prefix and '_DISA' suffix).
    - Moves processed ZIP files into a 'processed' subfolder for cleanliness.
    '''

    def __init__(self, source_folder: str, destination_folder: str):
        '''
        Initialize the NessusExtractor with source and destination folders.

        Args:
            source_folder (str): Path to the folder containing ZIP files.
            destination_folder (str): Path to the folder where extracted .nessus files will be saved.
        '''
        self.source_folder = Path(source_folder)
        self.destination_folder = Path(destination_folder)
        self.processed_folder = self.source_folder / "processed"

        # Ensure destination and processed folders exist
        self.destination_folder.mkdir(parents=True, exist_ok=True)
        self.processed_folder.mkdir(parents=True, exist_ok=True)
        print(f"⚙️  NessusExtractor initialized...")
        Logger.log(f"⚙️  NessusExtractor initialized...")

    def _friendly_name(self, zip_path: Path) -> str:
        '''
        Generate a clean, human-friendly filename from a ZIP archive.
        
        Args:
            zip_path (Path): Path to the ZIP file.
        
        Returns:
            str: The friendly name derived from the ZIP filename.
        '''
        friendly_name = zip_path.stem
        # Remove "PAAN_" prefix and "_DISA" suffix if present
        if friendly_name.startswith("PAAN_"):
            friendly_name = friendly_name[len("PAAN_"):]
        if friendly_name.endswith("_DISA"):
            friendly_name = friendly_name[:-len("_DISA")]
        return friendly_name + ".nessus"
    
    def _process_zip(self, zip_path: Path) -> List[Path]:
        '''
        Extract .nessus files from a ZIP archive.
        
        Args:
            zip_path (Path): Path to the ZIP file.
        
        Returns:
            List[Path]: List of paths to the extracted .nessus files.
        '''
        extracted_files = []
        friendly_name = self._friendly_name(zip_path)
        extracted_path = self.destination_folder / friendly_name
        
        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
            for file in zip_ref.namelist():
                if file.endswith(".nessus"):
                    with zip_ref.open(file) as src, open(extracted_path, "wb") as dst:
                        dst.write(src.read())
                    extracted_files.append(extracted_path)
                    print(f"[+] Extracted {file} from {zip_path.name} → {extracted_path}")
                    Logger.log(f"[+] Extracted {file} from {zip_path.name} → {extracted_path}")
                else:
                    print(f"[-] Skipping {file} (not a .nessus file)")
        
        # Move processed ZIP to "processed" folder
        # TEMPORARILY DISABLING THE MOVE
        shutil.move(str(zip_path), self.processed_folder / zip_path.name)
        print(f"[✓] Moved {zip_path.name} → {self.processed_folder.name}")
        Logger.log(f"[✓] Moved {zip_path.name} → {self.processed_folder.name}")
        
        return extracted_files

    def extract_all(self) -> List[Path]:
        '''
        Extract all .nessus files from ZIPs in the source folder.
        Rename them based on the ZIP filename.
        Returns a list of extracted file paths.
        '''
        print(f"⚙️  Starting Nessus Extraction...")
        Logger.log(f"⚙️  Starting Nessus Extraction...")

        all_extracted_files = []

        for zip_path in self.source_folder.glob("*.zip"):
            all_extracted_files.extend(self._process_zip(zip_path))

        print(f"✅ Nessus extraction complete.")
        Logger.log("✅ Nessus extraction complete.")
        return all_extracted_files
    

class NessusWorkflow:
    '''
    Orchestrates the Nessus workflow. Extraction, Parsing, and Exporting.
    '''

    def __init__(self, input_folder, output_file, cat_lvls=("II",)):
        self.input_folder = Path(input_folder).resolve()
        self.output_file = Path(output_file).resolve()
        self.cat_lvls = cat_lvls

    def run(self):
        print(f"📂 Starting Nessus workflow...")
        Logger.log(f"📂 Starting Nessus workflow...")

        # Always define extracted folder for subsequent runs of script
        extracted_folder = self.input_folder / "extracted_nessus"
        extracted_folder.mkdir(exist_ok=True)

        # Step 1: Extract .nessus files from ZIP files if required
        zip_files = list(self.input_folder.glob("*.zip"))
        if zip_files: #only extract if zips are present
            print(f"📦 Found {len(zip_files)} ZIP files to extract. Extracting from {self.input_folder}...")
            Logger.log(f"📦 Found {len(zip_files)} ZIP files to extract. Extracting from {self.input_folder}...")
            extractor = NessusExtractor(self.input_folder, extracted_folder)
            extractor.extract_all()
        else:
            print(f"ℹ️ No ZIP files found in {self.input_folder}, skipping extraction.")
            Logger.log(f"ℹ️ No ZIP files found in {self.input_folder}, skipping extraction.")

        # Step 2: Export CAT findings to Excel
        nessus_files = list(extracted_folder.glob("*.nessus"))
        print(f"📑 Processing {len(nessus_files)} Nessus files...")
        Logger.log(f"📑 Processing {len(nessus_files)} Nessus files...") 
        exporter = NessusToExcelExporter(extracted_folder, self.output_file, self.cat_lvls)
        exporter.run()

        print(f"✅ Finished processing Nessus files in {self.input_folder}")
        print(f"✅ Exported CAT findings")
        #print(f"✅ All done!")
        Logger.log(f"✅ All done! Finished processing Nessus files in {self.input_folder}, and exported CAT findings to {self.output_file}.")


class Logger:
    LOG_FILE = "log.txt"

    @staticmethod
    def log(message: str):
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        with open(Logger.LOG_FILE, "a", encoding="utf-8") as f:
            f.write(f"{timestamp} - {message}\n")
        #print(message)

def pick_folders_gui():
    root = tk.Tk()
    root.withdraw()

    input_folder = filedialog.askdirectory(
        title="Select the folder containing the Nessus XML files"
    )
    if not input_folder:
        print("❌ No folder selected. Exiting.")
        exit(1)

    output_file = filedialog.asksaveasfilename(
        title="Save Excel File As",
        defaultextension=".xlsx",
        filetypes=[("Excel Files", "*.xlsx")],
        initialfile="CATT_Extracted_Data.xlsx",
    )
    if not output_file:
        print("❌ No file selected. Exiting.")
        exit(1)
    
    return input_folder, output_file